import collections
import datetime
import functools
import io

import openpyxl
from django import forms
from django.conf import settings
from django.db.models import Exists
from django.db.models import OuterRef
from django.utils.decorators import method_decorator
from django.views.decorators.cache import never_cache
from django_filters import rest_framework as filters
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from rest_framework import viewsets
from rest_framework.mixins import ListModelMixin
from rest_framework.renderers import BaseRenderer

from digital_agenda.apps.charts.models import ChartGroup
from digital_agenda.apps.core.models import Breakdown
from digital_agenda.apps.core.models import BreakdownGroup
from digital_agenda.apps.core.models import Country
from digital_agenda.apps.core.models import Fact
from digital_agenda.apps.core.models import Indicator
from digital_agenda.apps.core.models import IndicatorGroup
from digital_agenda.apps.core.models import Period
from digital_agenda.apps.core.models import Unit
from digital_agenda.apps.core.serializers import FactSerializer
from digital_agenda.apps.core.views import DimensionViewSetMixin

EUROSTAT_FLAGS = {
    # Taken from https://ec.europa.eu/eurostat/data/database/information
    "b": "break in time series",
    "c": "confidential",
    "d": "definition differs, see metadata",
    "e": "estimated",
    "f": "forecast",
    "n": "not significant",
    "p": "provisional",
    "r": "revised",
    "s": "Eurostat estimate",
    "u": "low reliability",
    "z": "not applicable",
    # Custom flags, not from ESTAT; for internal use only
    "x": "unavailable",
    "~": "at least one datapoint missing from combined value",
}


class FactsFilter(filters.FilterSet):
    unit = filters.CharFilter(
        field_name="unit__code",
        label="unit.code",
        help_text="Filter results by unit code",
        required=True,
    )
    period = filters.CharFilter(
        field_name="period__code",
        label="period.code",
        help_text="Filter results by period " "code",
    )
    country = filters.CharFilter(
        field_name="country__code",
        label="country.code",
        help_text="Filter results by " "country code",
    )
    indicator = filters.CharFilter(
        field_name="indicator__code",
        label="indicator.code",
        help_text="Filter results by indicator code",
    )
    breakdown = filters.CharFilter(
        field_name="breakdown__code",
        label="breakdown.code",
        help_text="Filter results by breakdown code",
    )
    indicator_group = filters.CharFilter(
        field_name="indicator",
        label="indicator_group.code",
        help_text="Filter result by indicator_group code",
        method="filter_group",
    )
    breakdown_group = filters.CharFilter(
        field_name="breakdown",
        label="breakdown_group.code",
        help_text="Filter result by breakdown_group code",
        method="filter_group",
    )

    def filter_group(self, queryset, name, value):
        rel_model = getattr(queryset.model, name).field.related_model

        return queryset.filter(
            Exists(
                rel_model.objects.filter(id=OuterRef(f"{name}_id"), groups__code=value)
            )
        )

    def get_form_class(self):
        # Ensure that at least one indicator and one breakdown filter
        # has been used to avoid accidentally making huge queries.

        def clean_indicator(form):
            data = form.cleaned_data
            if not data["indicator_group"] and not data["indicator"]:
                raise forms.ValidationError(
                    "Either indicator or indicator_group is required"
                )
            return data["indicator"]

        def clean_breakdown(form):
            data = form.cleaned_data
            if not data["breakdown_group"] and not data["breakdown"]:
                raise forms.ValidationError(
                    "Either breakdown or breakdown_group is required"
                )
            return data["breakdown"]

        form_class = super().get_form_class()
        form_class.clean_indicator = clean_indicator
        form_class.clean_breakdown = clean_breakdown

        return form_class

    class Meta:
        model = Fact
        fields = [
            "indicator_group",
            "indicator",
            "breakdown_group",
            "breakdown",
            "unit",
            "period",
            "country",
        ]


MODELS = {
    "indicator": Indicator,
    "breakdown": Breakdown,
    "unit": Unit,
    "country": Country,
    "period": Period,
    "breakdown_group": BreakdownGroup,
    "indicator_group": IndicatorGroup,
}


class FactXLSXSerializer:
    def __init__(self, data, renderer_context):
        self.data = data
        self.view = renderer_context["view"]
        self.request = renderer_context["request"]
        self.headers = renderer_context["header"]
        self.wb = openpyxl.Workbook(write_only=True)

    def render(self):
        self.render_info_sheet()
        self.render_filters_sheet()
        self.render_dimensions_sheet()
        self.render_data_sheet()

        buffer = io.BytesIO()
        self.wb.save(buffer)

        yield buffer.getvalue()

    @functools.cached_property
    def chart_group(self):
        try:
            return ChartGroup.objects.get(code=self.request.GET.get("chart_group"))
        except ChartGroup.DoesNotExist:
            return None

    @functools.cached_property
    def filter_args(self):
        result = {}
        for filter_key in self.view.filterset_class.base_filters:
            if not (value := self.request.GET.get(filter_key)):
                continue

            result[filter_key] = value
        return result

    def set_dimensions(self, sheet, dimensions):
        for i, width in enumerate(dimensions, start=1):
            sheet.column_dimensions[get_column_letter(i)].width = width

    def write_headers(self, sheet, headers):
        cells = []
        for value in headers:
            cell = WriteOnlyCell(sheet, value)
            cell.font = Font(bold=True)
            cells.append(cell)
        sheet.append(cells)

    def render_info_sheet(self):
        host = settings.FRONTEND_HOST[0]

        sheet = self.wb.create_sheet("General Information")
        self.set_dimensions(sheet, [30, 100])
        sheet.append(["Extraction Date", str(datetime.date.today())])

        if not self.chart_group:
            return

        sheet.append(["Source", "European Commission, Digital Decade DESI visualisation tool"])
        sheet.append(["Dataset", self.chart_group.name])
        sheet.append(
            ["Charts", f"https://{host}/datasets/{self.chart_group.code}/charts"]
        )
        sheet.append(
            [
                "List of available indicators",
                f"https://{host}/datasets/{self.chart_group.code}/indicators",
            ]
        )
        sheet.append(
            [
                "Dataset metadata and download",
                f"https://{host}/datasets/{self.chart_group.code}/metadata",
            ]
        )

    def render_filters_sheet(self):
        sheet = self.wb.create_sheet("Applied Filters")
        self.set_dimensions(sheet, [15, 15, 40, 40, 40])
        self.write_headers(
            sheet, ["Dimension", "Code", "Label", "Alt. Label", "Definition"]
        )

        for filter_key, value in self.filter_args.items():
            obj = MODELS[filter_key].objects.get(code=value)

            sheet.append(
                [
                    self.chart_group.get_label(filter_key),
                    value,
                    obj.label,
                    obj.alt_label,
                    obj.definition,
                ]
            )

    def render_dimensions_sheet(self):
        headers = ["code", "label", "alt_label", "definition"]
        dimensions = ["indicator", "breakdown", "unit", "country", "period"]
        dimension_codes = collections.defaultdict(set)
        for row in self.data:
            for dim in dimensions:
                dimension_codes[dim].add(row[dim])

        for dim, codes in dimension_codes.items():
            if dim in self.filter_args:
                # If we are filtering by the dimension there is
                # no point in adding an extra sheet for it.
                continue

            model = MODELS[dim]

            sheet = self.wb.create_sheet(self.chart_group.get_label(dim))
            self.set_dimensions(sheet, [15, 40, 40, 40])
            self.write_headers(
                sheet, [getattr(model, h).field.verbose_name.title() for h in headers]
            )

            for obj in model.objects.filter(code__in=codes):
                sheet.append(getattr(obj, key) for key in headers)

        sheet = self.wb.create_sheet("Flags")
        self.set_dimensions(sheet, [10, 40])
        self.write_headers(sheet, ["Flag", "Label"])

        for key, value in EUROSTAT_FLAGS.items():
            sheet.append([key, value])

    def render_data_sheet(self):
        sheet = self.wb.create_sheet("Data")
        self.set_dimensions(sheet, [15 for _ in self.headers])
        self.write_headers(sheet, [self.chart_group.get_label(h) for h in self.headers])

        for row in self.data:
            sheet.append(row[header] for header in self.headers)


class FactsXLSXRenderer(BaseRenderer):
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    format = "xlsx"

    def render(self, data, accepted_media_type=None, renderer_context=None):
        yield from FactXLSXSerializer(data, renderer_context).render()


class FactsViewSet(DimensionViewSetMixin, ListModelMixin, viewsets.GenericViewSet):
    model = Fact
    serializer_class = FactSerializer
    filter_backends = [filters.DjangoFilterBackend]
    filterset_class = FactsFilter
    queryset = (
        Fact.objects.order_by(
            "period__code",
            "indicator__code",
            "breakdown__code",
            "unit__code",
            "country__code",
        )
        .select_related("country", "indicator", "breakdown", "period", "unit")
        .only(
            "period__code",
            "indicator__code",
            "breakdown__code",
            "unit__code",
            "country__code",
            "value",
            "flags",
        )
        .all()
    )
    renderer_classes = viewsets.GenericViewSet.renderer_classes + [FactsXLSXRenderer]

    # Likely to have a lot of cache misses, does not seem worth caching.
    @method_decorator(never_cache)
    def list(self, request, *args, **kwargs):
        """List observation values for the specified filters.
        Following filters are required:

         - `indicator_group` OR `indicator`
         - `breakdown_group` OR `breakdown`
         - `unit`

        XLSX format will also include sheets with the relevant metadata.

        """
        return super().list(request, *args, **kwargs)

    def get_renderer_context(self):
        return {
            **super().get_renderer_context(),
            "header": [
                "period",
                "country",
                "indicator",
                "breakdown",
                "unit",
                "value",
                "flags",
            ],
        }
