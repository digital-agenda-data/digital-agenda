import decimal
from abc import ABC, abstractmethod
from collections import defaultdict
import logging

from django.db import transaction
import xlrd
import openpyxl
from openpyxl.utils import get_column_letter

from .models import Period, Country, Indicator, Breakdown, Unit, Fact
from .views.facts import EUROSTAT_FLAGS

logger = logging.getLogger(__name__)

UNIQUE_EXCEL_COLS = (
    "period",
    "country",
    "indicator",
    "breakdown",
    "unit",
)
VALID_EXCEL_COLS = (
    "period",
    "country",
    "indicator",
    "breakdown",
    "unit",
    "value",
    "flags",
    "reference_period",
    "remarks",
)
DIMENSION_MODELS = {
    "indicator": Indicator,
    "breakdown": Breakdown,
    "unit": Unit,
    "country": Country,
    "period": Period,
}


def empty_cell(cell):
    return cell.value is None or cell.value == ""


class DimensionCache:
    def __init__(self, model):
        self.model = model
        self.cache = {}
        # Period obj can be auto-created
        self._auto_create = model == Period

    def get(self, code):
        try:
            return self.cache[code]
        except KeyError:
            pass

        if self._auto_create:
            obj = self.cache[code] = self.model.objects.get_or_create(code=code)[0]
        else:
            try:
                obj = self.cache[code] = self.model.objects.get(code=code)
            except (self.model.DoesNotExist, ValueError):
                obj = None

        return obj


class BaseFileLoader(ABC):
    """Base class for loading file data into database `Fact`s"""

    def __init__(self, path):
        self.path = path
        self.dimensions = {
            name: DimensionCache(model) for name, model in DIMENSION_MODELS.items()
        }

    @abstractmethod
    def load(self, *args, **kwargs): ...


class RowReader:
    def __init__(
        self,
        row,
        dimensions,
        header_columns,
        extra_fields=None,
    ):
        self.dimensions = dimensions
        self.header_columns = header_columns
        self.row = row
        self.row_dict = self._get_row_dict()
        self.errors = defaultdict(list)
        self.fields = {**extra_fields}
        self.read_row()

    def _get_row_dict(self):
        result = {}
        for field, index in self.header_columns.items():
            try:
                cell = self.row[index]
                assert not empty_cell(cell)
            except (IndexError, AssertionError):
                result[field] = None
                continue

            result[field] = cell.value
        return result

    def add_error(self, col_index, error):
        col_letter = get_column_letter(col_index + 1)
        self.errors[f"Column {col_letter}"].append(error)

    def get_value(self):
        value = self.row_dict["value"]

        if value is not None:
            try:
                # Excel can sometimes store values like 5.6621000000000006 while only
                # displaying 5.6621.
                # Round to a sensible precision to (hopefully) get rid of the issue.
                value = float(round(decimal.Decimal(value), 6))
            except (TypeError, ValueError, ArithmeticError):
                self.add_error(
                    self.header_columns["value"],
                    f"Invalid 'value', expected number but got {value!r} instead",
                )

        return value

    def get_flags(self):
        flags = self.row_dict.get("flags") or ""
        for flag in flags:
            if flag not in EUROSTAT_FLAGS:
                self.add_error(self.header_columns["flags"], f"Unknown flag {flag!r}")

        return flags

    def read_row(self):
        """
        Read data from the given row.
        """
        self.fields["value"] = self.get_value()
        self.fields["flags"] = self.get_flags()
        self.fields["reference_period"] = self.row_dict.get("reference_period")
        self.fields["remarks"] = self.row_dict.get("remarks")

        if self.fields["value"] is None and not self.fields["flags"]:
            # Set the custom flag "unavailable" for this case
            self.fields["flags"] = "x"

        for col in UNIQUE_EXCEL_COLS:
            if not self.row_dict.get(col):
                self.add_error(
                    self.header_columns[col], f"Column {col!r} must not be empty"
                )

        for dim_name, dim_model in DIMENSION_MODELS.items():
            dim_code = self.row_dict[dim_name]

            self.fields[dim_name] = self.dimensions[dim_name].get(dim_code)
            if self.fields[dim_name] is None:
                self.add_error(
                    self.header_columns[dim_name],
                    f"Missing dimension code for {dim_name!r}: {dim_code!r}",
                )


class BaseExcelLoader(BaseFileLoader, ABC):
    """
    Loader for Excel file formats.
    """

    def __init__(self, path, extra_fields=None):
        super().__init__(path)
        self.extra_fields = extra_fields or {}
        self.max_col_index = 0
        self.header_columns = {}
        self.sheet = None
        self.errors = {}
        self.read_headers()

    @property
    @abstractmethod
    def rows_iterator(self):
        """Returns an implementation-specific (xls/xlsx) iterator over the active
        sheet's rows.
        """
        ...

    @property
    @abstractmethod
    def header_row(self):
        """Returns an implementation-specific (xls/xlsx) header row for the
        active sheet.
        """
        ...

    @abstractmethod
    def get_row(self, row_ref):
        """Get a row object using a reference produced by `row_iterator`."""
        ...

    def read_headers(self):
        for index, cell in enumerate(self.header_row):
            if empty_cell(cell):
                break

            header_name = cell.value.lower().strip()
            assert header_name in VALID_EXCEL_COLS, f"{header_name} not valid"
            self.header_columns[header_name] = index
            self.max_col_index = index

    def read(self):
        """
        Read the all rows from Excel file.

        Collects Fact instances into the `data` attribute,
        and unknown dimension values into the `errors` attribute.
        """
        data = []
        errors = {}
        all_unique_keys = {}

        for row_index, row_ref in enumerate(self.rows_iterator, start=1):
            row_reader = RowReader(
                self.get_row(row_ref),
                self.dimensions,
                header_columns=self.header_columns,
                extra_fields=self.extra_fields,
            )

            unique_key = tuple(row_reader.fields[field] for field in UNIQUE_EXCEL_COLS)
            if duplicate_row := all_unique_keys.get(unique_key):
                row_reader.errors["ALL"].append(
                    f"Duplicate entry found at Row {duplicate_row}"
                )
            all_unique_keys[unique_key] = row_index

            # Skip the row if any dimension code is unknown or both value and flags are missing
            if row_reader.errors:
                errors[f"Row {row_index}"] = dict(row_reader.errors)
                continue

            data.append(Fact(**row_reader.fields))

        return data, errors

    def load(self, allow_errors=False):
        """
        Load the Excel data into `Fact` records.
        If errors are encountered, data is written to the database only if `allow_errors` is set.

        Returns: tuple(int, dict)
            - number of records loaded
            - errors
        """
        data, errors = self.read()

        if not data and not errors:
            return 0, {"issue": "No valid row found"}
        elif errors and not allow_errors:
            return 0, errors

        with transaction.atomic():
            facts = Fact.objects.bulk_create(
                data,
                update_conflicts=True,
                update_fields=(
                    "value",
                    "flags",
                    "import_file",
                    "reference_period",
                    "remarks",
                ),
                unique_fields=(
                    "indicator",
                    "breakdown",
                    "unit",
                    "country",
                    "period",
                ),
            )
            return len(facts), errors


class XLSLoader(BaseExcelLoader):
    def __init__(self, path, *args, **kwargs):
        self.wb = xlrd.open_workbook(path)
        self.ws = self.wb.sheet_by_index(0)
        super().__init__(path, *args, **kwargs)

    def get_row(self, row_ref):
        return self.ws.row(row_ref)

    @property
    def header_row(self):
        return self.ws.row(0)

    @property
    def rows_iterator(self):
        return range(1, self.ws.nrows)


class XLSXLoader(BaseExcelLoader):
    def __init__(self, path, *args, **kwargs):
        self.wb = openpyxl.load_workbook(path, read_only=True)
        self.ws = self.wb.active
        super().__init__(path, *args, **kwargs)

    def get_row(self, row_ref):
        return row_ref  # openpyxl iter_rows returns actual row objects

    @property
    def header_row(self):
        return next(self.ws.iter_rows())

    @property
    def rows_iterator(self):
        return self.ws.iter_rows(2, self.ws.max_row, 1, self.max_col_index + 1)


def get_loader(data_file, extra_fields=None):
    if data_file.mime_type == "application/vnd.ms-excel":
        return XLSLoader(data_file.path, extra_fields=extra_fields)
    elif (
        data_file.mime_type
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ):
        return XLSXLoader(data_file.path, extra_fields=extra_fields)
    else:
        raise TypeError("Unsupported MIME type")
